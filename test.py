# coding:utf-8
import sys
import pandas as pd
from PyQt6.QtWidgets import QToolTip,QPushButton,QMenu,QFrame,QMessageBox
from PyQt6.QtGui import QColor, QPixmap, QIcon
from PyQt6.QtWidgets import QApplication,QTableWidget
from qframelesswindow import AcrylicWindow,TitleBar,StandardTitleBar
from window import Ui_MainWindow
from PyQt6.QtWidgets import QApplication, QMainWindow
from PyQt6.QtGui import QFont
from PyQt6 import QtCore, QtGui, QtWidgets
from openpyxl import load_workbook
data_path=r"./data/data_1.xlsx"
data = pd.read_excel(data_path)
row_count = data.shape[0]+1
line_count = data.shape[1]+1
horizontalHeader = ["时间","收入","支出"]
class Window(AcrylicWindow,QMainWindow, Ui_MainWindow):
    def settings(self):
        QMessageBox.about(self, 'Setting','你来这想干嘛呢🤣🤣🤣')
    def about(self):
        QMessageBox.about(self, '关于', """Version:1.0.0\nAuthor:Yuhongyi""")
    def homein(self):
        self.tableWidget.setColumnHidden(2,True)#删除指定列
    def homeout(self):
        self.tableWidget.setColumnHidden(1,True)
    def all(self):
        self.tableWidget.setColumnHidden(2,False)
        self.tableWidget.setColumnHidden(1,False)
    def __init__(self, parent=None):
        super().__init__(parent=parent)
        self.titleBar.raise_()
        self.setTitleBar(StandardTitleBar(self))
        self.setWindowIcon(QIcon("src/—Pngtree—pig piggy bank_4440424.png"))#设置图表
        self.setWindowTitle("Home bank")#设置标题
        self.updateFrameless()
        #创建组件
        Frame1=QFrame(self)
        Frame1.move(20,40)
        Frame1.resize(131,921)
        Frame1.setStyleSheet("background-color: rgba(255, 255, 255, 0.8);border-radius: 15px")
        Frame2=QFrame(self)
        Frame2.move(170,40)
        Frame2.resize(1181,921)
        Frame2.setStyleSheet("background-color: rgba(255, 255, 255, 0.8);border-radius: 15px")
        self.tableWidget = QTableWidget(self)

        # 读取Excel文件
        wb = load_workbook('./data/data_1.xlsx')
        sheet = wb.active
        # 获取行数和列数
        rows = sheet.max_row
        cols = sheet.max_column
        # 设置表格的行数和列数
        self.tableWidget.setRowCount(rows)
        self.tableWidget.setColumnCount(cols)
        self.tableWidget.move(180,60)
        self.tableWidget.resize(1161,881)
        # 读取数据并显示在表格中
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                cell_value = sheet.cell(row=row, column=col).value
                item =QtWidgets.QTableWidgetItem(str(cell_value))
                self.tableWidget.setItem(row - 1, col - 1, item)
        Button1=QPushButton("综合",self)
        Button1.setToolTip("查看家庭此月综合收入/支出情况")
        Button1.move(30,60)
        Button1.resize(111,51)
        Button1.clicked.connect(self.all)
        Button1.setStyleSheet("""
                                QPushButton {
                                    border-radius: 10px;
                                    border:10px;
                                    background-color: rgb(255,255,255);
                                }
                                QPushButton::hover {
	                                background-color: rgba(103, 216, 217,150);
                                }
                                   """)
        Button2=QPushButton("收入",self)
        Button2.setToolTip("查看家庭此月收入情况")
        Button2.move(30,120)
        Button2.resize(111,51)
        Button2.clicked.connect(self.homein)
        Button2.setStyleSheet("""
                                QPushButton {
                                    border-radius: 10px;
                                    border:10px;
                                    background-color: rgb(255,255,255);
                                }
                                QPushButton::hover {
	                                background-color: rgba(103, 216, 217,150);
                                }
                                   """)
        Button3=QPushButton("支出",self)
        Button3.setToolTip("查看家庭此月支出情况")
        Button3.move(30,180)
        Button3.resize(111,51)
        Button3.clicked.connect(self.homeout)
        Button3.setStyleSheet("""
                                QPushButton {
                                    border-radius: 10px;
                                    border:10px;
                                    background-color: rgb(255,255,255);
                                }
                                QPushButton::hover {
	                                background-color: rgba(103, 216, 217,150);
                                }
                                   """)
        Button4=QPushButton("设置",self)
        Button4.setToolTip("更改软件设置")
        Button4.move(30,901)
        Button4.resize(111,51)
        Button4.clicked.connect(self.settings)
        Button4.setStyleSheet("""
                                QPushButton {
                                    border-radius: 10px;
                                    border:10px;
                                    background-color: rgb(255,255,255);
                                }
                                QPushButton::hover {
	                                background-color: rgba(103, 216, 217,150);
                                }
                                   """)
        Button5=QPushButton("关于",self)
        Button5.setToolTip("查看软件相关")
        Button5.move(30,841)
        Button5.resize(111,51)
        Button5.clicked.connect(self.about)
        Button5.setStyleSheet("""
                                QPushButton {
                                    border-radius: 10px;
                                    border:10px;
                                    background-color: rgb(255,255,255);
                                }
                                QPushButton::hover {
	                                background-color: rgba(103, 216, 217,150);
                                }
                                   """)
        

        # customize acrylic effect
        # self.windowEffect.setAcrylicEffect(self.winId(), "106EBE99")
        # self.windowEffect.setAeroEffect(self.winId())#设置Aero模糊
        # you can also enable mica effect on Win11
        # self.windowEffect.setMicaEffect(self.winId(), isDarkMode=True, isAlt=False)
        
class CustomTitleBar(StandardTitleBar):#设置标题栏
    """ Custom title bar """

    def __init__(self, parent):
        super().__init__(parent)
        # use qss to customize title bar button
        self.maxBtn.setStyleSheet("""
            TitleBarButton {
                qproperty-hoverColor: white;
                qproperty-hoverBackgroundColor: rgb(0, 100, 182);
                qproperty-pressedColor: white;
                qproperty-pressedBackgroundColor: rgb(54, 57, 65);
            }
        """)
if __name__ == '__main__':
    app = QApplication(sys.argv)
    demo = Window()
    demo.resize(1364, 977)
    demo.setFixedSize(demo.width(), demo.height());    
    demo.show()
    sys.exit(app.exec())